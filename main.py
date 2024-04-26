# import libraries
import sys
from PySide6 import QtWidgets
from PySide6.QtWidgets import QWidget, QWidgetAction, QApplication, QMainWindow, QMdiArea, QMdiSubWindow, QTableWidget, \
    QFileDialog, QMessageBox, QTableWidgetItem
from PySide6.QtGui import Qt
from datetime import datetime, date
from Conexion import ConexionMySQL
import mysql.connector
from openpyxl import load_workbook

# import guis and libraries
from inter.MainWindow import Ui_MainWindow
from inter.juntas import *
from inter.instituciones import *
from inter.Hogar import *
from inter.planesInversion import *


# Mysql
def inJun(PJ, fecVen, dis, ubi, tele, dirNom, dirEmail):
    aux = "INSERT INTO Juntas (PJ, fechVen, distri, ubicacion, telefono, dirNom, dirEmail) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    datos = (PJ, fecVen, dis, ubi, tele, dirNom, dirEmail)
    #start conexion
    conexionMSQL = ConexionMySQL("192.168.100.100", "kelunie", "culep1711", "proyecto")
    conexionMSQL.conectar()

    cursor = conexionMSQL.conexion.cursor()
    try:
        # Ejecutar la consulta SQL
        cursor.execute(aux, datos)

        # Confirmar los cambios
        conexionMSQL.conexion.commit()

        print("Datos insertados correctamente en la tabla Juntas.")
    except Exception as e:
        print(f"Error al insertar datos: {e}")
        # Deshacer la transacción en caso de error
        conexionMSQL.conexion.rollback()

    # Cerrar el cursor y la conexión
    cursor.close()
    conexionMSQL.desconectar()


def ininst(nom, PJ, vigen, func, dis, ubi, tele, email, presidente, tesorero):
    try:
        # Establecer conexión
        conexionMSQL = ConexionMySQL("192.168.100.100", "kelunie", "culep1711", "proyecto")
        conexionMSQL.conectar()
        cursor = conexionMSQL.conexion.cursor()

        # Insertar datos de la Institución
        aux = "INSERT INTO Instituciones (nombre, personeriaJuridica, vigencia, funciones, distrito, ubicacion, telefono, correo) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
        datos = (nom, PJ, vigen, func, dis, ubi, tele, email)
        cursor.execute(aux, datos)

        # Recuperar el ID de la Institución
        auxconsult = "SELECT idInstitucion FROM Instituciones WHERE personeriaJuridica = %s"
        cursor.execute(auxconsult, (PJ,))
        id_institucion = cursor.fetchone()

        if id_institucion:
            # Insertar datos de la Junta Administrativa
            aux2 = "INSERT INTO JuntaAdministrativa (idInstitucion, presidente, tesorero) VALUES (%s, %s, %s)"
            datos2 = (id_institucion[0], presidente, tesorero)
            cursor.execute(aux2, datos2)

        # Confirmar los cambios
        conexionMSQL.conexion.commit()
        print("Datos insertados correctamente en las tablas Instituciones y JuntaAdministrativa.")

    except mysql.connector.Error as e:
        print(f"Error al insertar datos: {e}")
        # Deshacer la transacción en caso de error
        conexionMSQL.conexion.rollback()

    finally:
        # Cerrar cursor y conexión
        if cursor:
            cursor.close()
        if conexionMSQL:
            conexionMSQL.desconectar()


def inhogar(nomb, personeria_juridica, tiempo_vigencia, distrito, ubicacion, telefono, correo, tipo_centro, valoracion_puntos,
            horario_atencion, poblacion_anual, fecha_entrega_plan, fecha_presentacion_informe, presidente, tesorero,
            banco, cuenta_corriente):
    try:
        # Establecer conexión
        conexionMSQL = ConexionMySQL("192.168.100.100", "kelunie", "culep1711", "proyecto")
        conexionMSQL.conectar()
        cursor = conexionMSQL.conexion.cursor()

        # Insertar datos del Hogar
        aux = "INSERT INTO Hogares (nombre, personeria_juridica, tiempo_vigencia, distrito, ubicacion, telefono, correo, tipo_centro, valoracion_puntos, horario_atencion, poblacion_anual, fecha_entrega_plan, fecha_presentacion_informe) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        datos = (
            nomb, personeria_juridica, tiempo_vigencia, distrito, ubicacion, telefono, correo, tipo_centro, valoracion_puntos,
            horario_atencion, poblacion_anual, fecha_entrega_plan, fecha_presentacion_informe)
        cursor.execute(aux, datos)

        # Recuperar el ID del Hogar
        consult = "SELECT id FROM Hogares WHERE `personeria_juridica` = %s"
        cursor.execute(consult, (personeria_juridica,))
        hogar_id = cursor.fetchone()

        if hogar_id:
            hogar_id = hogar_id[0]  # Convertir a un solo valor

            # Insertar datos de la Junta Directiva
            aux2 = "INSERT INTO JuntaDirectiva (hogar_id, presidente, tesorero) VALUES (%s, %s, %s)"
            datos2 = (hogar_id, presidente, tesorero)
            cursor.execute(aux2, datos2)


            # Insertar datos de las cuentas bancarias
            aux3 = "INSERT INTO CuentasBancarias (hogar_id, banco, cuenta_corriente) VALUES (%s, %s, %s)"
            datos_cuentas = (hogar_id, banco, cuenta_corriente)
            # Insertar datos de las cuentas bancarias
            print("Consulta SQL para cuentas bancarias:", aux3)
            print("Datos de cuentas bancarias:", datos_cuentas)
            cursor.execute(aux3, datos_cuentas)

        # Confirmar los cambios
        conexionMSQL.conexion.commit()
        print("Datos insertados correctamente en las tablas Hogares, JuntaDirectiva y CuentasBancarias.")

    except mysql.connector.Error as e:
        print(f"Error al insertar datos: {e}")
        # Deshacer la transacción en caso de error
        conexionMSQL.conexion.rollback()

    finally:
        # Cerrar cursor y conexión
        if cursor:
            cursor.close()
        if conexionMSQL:
            conexionMSQL.desconectar()


class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.ndi = QMdiArea()
        self.setCentralWidget(self.ndi)

        self.subRJuntas = QMdiSubWindow()
        self.subRInstituciones = QMdiSubWindow()
        self.subRHogares = QMdiSubWindow()
        self.subPlanesInversion = QMdiSubWindow()

        self.initMenu()

    def initMenu(self):
        # Conectar acciones del menú a funciones
        self.ui.actionJuntas.triggered.connect(self.abrirJuntas)
        self.ui.actionInstituciones.triggered.connect(self.abrirInstituciones)
        self.ui.actionHogares.triggered.connect(self.abrirHogares)
        self.ui.actionPlanes_Inversi_n.triggered.connect(self.abrirPlanesInversion)
        self.ui.actionInforme_Liquidaci_n.triggered.connect(self.abrirInformeLiquidacion)
        self.ui.actionSalir.triggered.connect(self.salir)

    # Funciones para abrir otras ventanas
    def abrirJuntas(self):
        self.RJuntas = Ui_Form_juntas()
        self.subRJuntas.setWidget(self.RJuntas)
        self.subRJuntas.setAttribute(Qt.WA_DeleteOnClose, False)
        self.ndi.addSubWindow(self.subRJuntas)
        self.RJuntas.btncancelar.clicked.connect(self.cerrarRJuntas)
        self.subRJuntas.setFixedSize(self.RJuntas.size())

        self.subRJuntas.show()
        self.RJuntas.btnIngresar.clicked.connect(self.ingresarJuntas)

    def ingresarJuntas(self):
        # Obtener los datos de los campos de entrada
        PJ = self.RJuntas.insPJ.text()
        fecVen = self.RJuntas.insFV.text()
        dis = self.RJuntas.insDis.text()
        ubi = self.RJuntas.insUbi.text()
        tele = int(self.RJuntas.insTel.text())
        dirNom = self.RJuntas.insED.text()
        dirEmail = self.RJuntas.insED.text()

        try:
            inJun(PJ, fecVen, dis, ubi, tele, dirNom, dirEmail)

            QMessageBox.information(self, "Éxito", "Los datos se han ingresado correctamente.")

            # clr all fields
            self.RJuntas.insPJ.setText("")
            self.RJuntas.insFV.setText("")
            self.RJuntas.insDis.setText("")
            self.RJuntas.insUbi.setText("")
            self.RJuntas.insND.setText("")
            self.RJuntas.insED.setText("")
            self.RJuntas.insTel.setText("")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al ingresar los datos: {str(e)}")

    def cerrarRJuntas(self):
        self.RJuntas.close()
        self.subRJuntas.hide()

    def cerrarRInstituciones(self):
        self.RInstituciones.close()
        self.subRInstituciones.hide()

    def cerrarRHogares(self):
        self.RHogares.close()
        self.subRHogares.hide()
    def cerrarPlanesIinversion(self):
        self.APlanesInversion.close()
        self.subPlanesInversion.hide()

    def abrirInstituciones(self):
        self.RInstituciones = Ui_instituciones()
        self.subRInstituciones.setWidget(self.RInstituciones)
        self.subRInstituciones.setAttribute(Qt.WA_DeleteOnClose, False)
        self.ndi.addSubWindow(self.subRInstituciones)
        self.RInstituciones.btncancelarins.clicked.connect(self.cerrarRInstituciones)
        self.subRInstituciones.setFixedSize(self.RInstituciones.size())

        self.subRInstituciones.show()
        self.RInstituciones.btningreins.clicked.connect(self.ingresarInstituciones)

    def ingresarInstituciones(self):
        nom = self.RInstituciones.innom.text()
        PJ = self.RInstituciones.inperju.text()
        vigen = self.RInstituciones.invige.text()
        func = self.RInstituciones.insfunciones.toPlainText()
        dis = self.RInstituciones.indis.text()
        ubi = self.RInstituciones.inubi.text()
        tele = int(self.RInstituciones.intel.text())
        email = self.RInstituciones.incorreo.text()
        presi = self.RInstituciones.inpresi.text()
        teso = self.RInstituciones.intesorero.text()

        try:
            ininst(nom, PJ, vigen, func, dis, ubi, tele, email, presi, teso)
            QMessageBox.information(self, "Éxito!", "Los datos se han ingresado correctamente.")

            # clr all fields
            self.RInstituciones.innom.setText("")
            self.RInstituciones.inperju.setText("")
            self.RInstituciones.invige.setText("")
            self.RInstituciones.insfunciones.setText("")
            self.RInstituciones.indis.setText("")
            self.RInstituciones.inubi.setText("")
            self.RInstituciones.intel.setText("")
            self.RInstituciones.incorreo.setText("")
            self.RInstituciones.inpresi.setText("")
            self.RInstituciones.intesorero.setText("")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al ingresar los datos: {str(e)}")

    def abrirHogares(self):
        self.RHogares = Ui_Hogar()
        self.subRHogares.setWidget(self.RHogares)
        self.subRHogares.setAttribute(Qt.WA_DeleteOnClose, False)
        self.ndi.addSubWindow(self.subRHogares)
        self.RHogares.btncancelar.clicked.connect(self.cerrarRHogares)
        self.subRHogares.setFixedSize(self.RHogares.size())

        self.subRHogares.show()
        self.RHogares.btningresar.clicked.connect(self.ingresarHogares)

    def ingresarHogares(self):
        nom = self.RHogares.insnomb.text()
        PJ = self.RHogares.insPJ.text()
        TV = self.RHogares.insTV.text()
        dis = self.RHogares.insdis.text()
        ubi = self.RHogares.insubi.text()
        tel = int(self.RHogares.instel.text())
        correo = self.RHogares.inscorreo.text()
        TC = self.RHogares.insTC.text()
        val = int(self.RHogares.insVP.text())
        HA = self.RHogares.insHA.text()
        PA = self.RHogares.insPA.text()
        FEP = self.RHogares.insFEP.text()
        FPI = self.RHogares.insFPI.text()
        pres = self.RHogares.inspresi.text()
        teso = self.RHogares.insteso.text()
        ban = self.RHogares.insbanco.text()
        CC = self.RHogares.insCC.text()

        try:
            inhogar(nom, PJ, TV, dis, ubi, tel, correo, TC, val, HA, PA, FEP, FPI, pres, teso, ban, CC)
            QMessageBox.information(self, "Éxito!", "Los datos se han ingresado Correctamente")

            # clr fields
            self.RHogares.insnomb.setText("")
            self.RHogares.insPJ.setText("")
            self.RHogares.insTV.setText("")
            self.RHogares.insdis.setText("")
            self.RHogares.insubi.setText("")
            self.RHogares.instel.setText("")
            self.RHogares.inscorreo.setText("")
            self.RHogares.insTC.setText("")
            self.RHogares.insVP.setText("")
            self.RHogares.insHA.setText("")
            self.RHogares.insPA.setText("")
            self.RHogares.insFEP.setText("")
            self.RHogares.insFPI.setText("")
            self.RHogares.inspresi.setText("")
            self.RHogares.insteso.setText("")
            self.RHogares.insbanco.setText("")
            self.RHogares.insCC.setText("")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al ingresar los datos: {str(e)}")

    def abrirPlanesInversion(self):
        self.APlanesInversion = Ui_planInversion()
        self.subPlanesInversion.setWidget(self.APlanesInversion)
        self.subPlanesInversion.setAttribute(Qt.WA_DeleteOnClose, False)
        self.ndi.addSubWindow(self.subPlanesInversion)
        self.APlanesInversion.btncloose.clicked.connect(self.cerrarPlanesIinversion)
        self.subPlanesInversion.setFixedSize(self.APlanesInversion.size())

        self.subPlanesInversion.show()
        self.APlanesInversion.btnopen.clicked.connect(self.abrir_archivo)
        self.APlanesInversion.btnload.clicked.connect(self.cargar_archivo)

    def abrir_archivo(self):
        file = QFileDialog.getOpenFileName(self, "Abrir Archivo Excel", "", "Excel files (*.xlsx) ;; All Files(*)")
        self.direccion = file[0]

    def cargar_archivo(self):
        try:
            # Verificar si se seleccionó un archivo
            if not self.direccion:
                QMessageBox.warning(self, "Advertencia", "No se seleccionó ningún archivo.")
                return

            # Cargar el libro de trabajo Excel
            libro_excel = load_workbook(self.direccion)

            # Leer la primera hoja de trabajo
            hoja = libro_excel.active

            # Leer datos de la hoja de trabajo
            datos = []

            total = 0

            # Leer filas y columnas
            for fila in hoja.iter_rows(values_only=True):
                fila_datos = []
                for valor in fila:
                    if valor is not None:  # Verificar si el valor no es None
                        fila_datos.append(valor)
                if fila_datos:  # Agregar solo si la fila contiene datos
                    datos.append(fila_datos)

            # add fields data
            self.APlanesInversion.insMuni.setText(datos[0][0])
            self.APlanesInversion.insSoli.setText(datos[1][0])
            self.APlanesInversion.insLey.setText(datos[2][0])
            # Agregar datos a la tabla table1
            self.APlanesInversion.table1.setItem(0, 0, QTableWidgetItem(datos[3][1]))
            self.APlanesInversion.table1.setItem(1, 0, QTableWidgetItem(datos[4][1]))
            self.APlanesInversion.table1.setItem(2, 0, QTableWidgetItem(datos[5][1]))
            self.APlanesInversion.table1.setItem(3, 0, QTableWidgetItem(datos[6][1]))
            self.APlanesInversion.table1.setItem(4, 0, QTableWidgetItem(datos[7][1]))

            # Agregar datos a la tabla table2
            self.APlanesInversion.table2.insertRow(0)
            for i in range(len(datos[9])):
                item = QTableWidgetItem(str(datos[9][i]))
                self.APlanesInversion.table2.setItem(0, i, item)
                if i == 5:
                    cant = int(datos[9][3])
                    PUni = int(datos[9][4])
                    subtotal = cant * PUni
                    total += subtotal
                    subtotal_item = QTableWidgetItem(str(subtotal))
                    self.APlanesInversion.table2.setItem(0, 5, subtotal_item)

            self.APlanesInversion.table2.insertRow(1)
            for i in range(len(datos[10])):
                item = QTableWidgetItem(str(datos[10][i]))
                self.APlanesInversion.table2.setItem(1, i, item)
                if i == 5:
                    cant = int(datos[10][3])
                    PUni = int(datos[10][4])
                    subtotal = cant * PUni
                    total += subtotal
                    subtotal_item = QTableWidgetItem(str(subtotal))
                    self.APlanesInversion.table2.setItem(1, 5, subtotal_item)

            self.APlanesInversion.table2.insertRow(2)
            for i in range(len(datos[11])):
                item = QTableWidgetItem(str(datos[11][i]))
                self.APlanesInversion.table2.setItem(2, i, item)
                if i == 5:
                    cant = int(datos[11][3])
                    PUni = int(datos[11][4])
                    subtotal = cant * PUni
                    total += subtotal
                    subtotal_item = QTableWidgetItem(str(subtotal))
                    self.APlanesInversion.table2.setItem(2, 5, subtotal_item)

            self.APlanesInversion.table2.insertRow(3)  # Agregar una nueva fila en el índice 3 (cuarta fila)
            for i in range(len(datos[12])):
                item = QTableWidgetItem(str(datos[12][i]))
                self.APlanesInversion.table2.setItem(3, i,
                                                     item)  # Agregar los elementos de datos[12] en la cuarta fila de table2
                if i == 5:
                    cant = int(datos[12][3])
                    PUni = int(datos[12][4])
                    subtotal = cant * PUni
                    total += subtotal
                    subtotal_item = QTableWidgetItem(str(subtotal))
                    self.APlanesInversion.table2.setItem(3, 5,
                                                         subtotal_item)  # Agregar el subtotal en la cuarta fila, columna 5

            total_str = str(total)
            self.APlanesInversion.instotal.setText(total_str)

            # Ajustar el tamaño de las columnas de table1
            for columna in range(self.APlanesInversion.table1.columnCount()):
                self.APlanesInversion.table1.resizeColumnToContents(columna)

            QMessageBox.information(self, "Información", "Archivo Excel cargado correctamente.")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al cargar el archivo Excel: {str(e)}")

    def abrirInformeLiquidacion(self):
        # Lógica para abrir la ventana de Informe de Liquidación
        pass

    def salir(self):
        # Lógica para salir de la aplicación
        pass


if __name__ == "__main__":
    app = QApplication(sys.argv)
    ventana_principal = MainWindow()
    ventana_principal.show()
    sys.exit(app.exec())
