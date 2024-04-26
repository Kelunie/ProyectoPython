from openpyxl import load_workbook
import os

def leer_excel(ruta_archivo):
    # Cargar el libro de trabajo
    libro_excel = load_workbook(ruta_archivo)

    # Seleccionar la primera hoja de trabajo
    hoja = libro_excel.active

    # Leer datos de la hoja de trabajo
    datos = []

    # Leer filas y columnas
    for fila in hoja.iter_rows(values_only=True):
        fila_datos = []
        for valor in fila:
            if valor is not None:  # Verificar si el valor no es None
                fila_datos.append(valor)
        if fila_datos:  # Agregar solo si la fila contiene datos
            datos.append(fila_datos)

    return datos

# Nombre de archivo de entrada y ruta completa
nombre_archivo = "Plan_Inversión_Casa_Cultura_2022-2023.xlsx"
ruta_completa = os.path.join("sources", nombre_archivo)

# Leer datos del archivo de entrada
datos_leidos = leer_excel(ruta_completa)

# Imprimir los datos leídos
print(datos_leidos[13][1])

